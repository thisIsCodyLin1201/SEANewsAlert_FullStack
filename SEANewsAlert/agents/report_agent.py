"""
Report Generator Agent
負責將 Markdown 報告轉換為 PDF 和 Excel
"""
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict
import markdown
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from config import Config
import re
from html.parser import HTMLParser
import pandas as pd


class HTMLToTextParser(HTMLParser):
    """簡單的 HTML 轉文字解析器"""
    def __init__(self):
        super().__init__()
        self.text = []
        
    def handle_data(self, data):
        self.text.append(data)
        
    def get_text(self):
        return ''.join(self.text)


class ReportGeneratorAgent:
    """報告生成代理 - 將 Markdown 轉換為專業 PDF"""
    
    def __init__(self):
        """初始化 Report Generator Agent"""
        self.reports_dir = Config.REPORTS_DIR
        self.setup_styles()
        
    def setup_styles(self):
        """設置 PDF 樣式"""
        self.styles = getSampleStyleSheet()
        
        # 嘗試註冊中文字體（支援 Windows 和 macOS）
        try:
            # Windows 和 macOS 常見中文字體路徑
            font_paths = [
                # Windows 系統
                'C:\\Windows\\Fonts\\msjh.ttc',           # 微軟正黑體
                'C:\\Windows\\Fonts\\msyh.ttc',           # 微軟雅黑
                'C:\\Windows\\Fonts\\kaiu.ttf',           # 標楷體
                'C:\\Windows\\Fonts\\mingliu.ttc',        # 細明體
                # macOS 系統
                '/System/Library/Fonts/PingFang.ttc',     # 蘋方（macOS 預設）
                '/System/Library/Fonts/STHeiti Light.ttc', # 華文黑體
                '/System/Library/Fonts/STHeiti Medium.ttc',
                '/Library/Fonts/Songti.ttc',              # 宋體
                '/System/Library/Fonts/Hiragino Sans GB.ttc', # 冬青黑體
                # Linux 系統（額外支援）
                '/usr/share/fonts/truetype/arphic/uming.ttc',
                '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
            ]
            
            font_registered = False
            for font_path in font_paths:
                if Path(font_path).exists():
                    try:
                        # 註冊字體
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                        self.chinese_font = 'ChineseFont'
                        font_registered = True
                        print(f"✅ 已註冊中文字體: {font_path}")
                        break
                    except Exception as font_error:
                        # 某些字體文件可能無法載入，繼續嘗試下一個
                        continue
            
            if not font_registered:
                print("⚠️  未找到中文字體，使用預設字體（可能無法正確顯示中文）")
                self.chinese_font = 'Helvetica'
        except Exception as e:
            print(f"⚠️  字體註冊失敗，使用預設字體: {str(e)}")
            self.chinese_font = 'Helvetica'
        
        # 標題樣式
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor='#1a5490',
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=self.chinese_font if hasattr(self, 'chinese_font') else 'Helvetica-Bold'
        ))
        
        # 副標題樣式
        self.styles.add(ParagraphStyle(
            name='CustomHeading2',
            parent=self.styles['Heading2'],
            fontSize=18,
            textColor='#2c5aa0',
            spaceAfter=12,
            spaceBefore=12,
            fontName=self.chinese_font if hasattr(self, 'chinese_font') else 'Helvetica-Bold'
        ))
        
        # 三級標題樣式
        self.styles.add(ParagraphStyle(
            name='CustomHeading3',
            parent=self.styles['Heading3'],
            fontSize=14,
            textColor='#3d6bb3',
            spaceAfter=8,
            spaceBefore=8,
            fontName=self.chinese_font if hasattr(self, 'chinese_font') else 'Helvetica-Bold'
        ))
        
        # 正文樣式
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['BodyText'],
            fontSize=11,
            leading=16,
            alignment=TA_JUSTIFY,
            fontName=self.chinese_font if hasattr(self, 'chinese_font') else 'Helvetica'
        ))
    
    def generate_pdf(
        self, 
        markdown_content: str, 
        filename: Optional[str] = None
    ) -> Path:
        """
        生成 PDF 報告
        
        Args:
            markdown_content: Markdown 格式的報告內容
            filename: 可選的文件名，不提供則自動生成
            
        Returns:
            Path: 生成的 PDF 文件路徑
        """
        print("📄 Report Generator Agent 開始生成 PDF...")
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"東南亞金融新聞報告_{timestamp}.pdf"
        
        # 確保文件名以 .pdf 結尾
        if not filename.endswith('.pdf'):
            filename += '.pdf'
        
        pdf_path = self.reports_dir / filename
        
        try:
            # 創建 PDF 文檔
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # 解析 Markdown 並生成內容
            story = self._parse_markdown_to_story(markdown_content)
            
            # 生成 PDF
            doc.build(story)
            
            print(f"✅ PDF 生成成功: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            print(f"❌ PDF 生成失敗: {str(e)}")
            raise
    
    def _parse_markdown_to_story(self, markdown_content: str):
        """將 Markdown 內容轉換為 ReportLab Story"""
        story = []
        lines = markdown_content.split('\n')
        
        for line in lines:
            line = line.strip()
            
            if not line:
                story.append(Spacer(1, 0.2*inch))
                continue
            
            # H1 標題
            if line.startswith('# '):
                text = line[2:].strip()
                story.append(Paragraph(text, self.styles['CustomTitle']))
                story.append(Spacer(1, 0.3*inch))
            
            # H2 標題
            elif line.startswith('## '):
                text = line[3:].strip()
                story.append(Paragraph(text, self.styles['CustomHeading2']))
                story.append(Spacer(1, 0.2*inch))
            
            # H3 標題
            elif line.startswith('### '):
                text = line[4:].strip()
                story.append(Paragraph(text, self.styles['CustomHeading3']))
                story.append(Spacer(1, 0.1*inch))
            
            # 列表項
            elif line.startswith('- ') or line.startswith('* '):
                text = '• ' + line[2:].strip()
                # 清理 Markdown 格式（包含粗體、斜體和連結）
                text = self._clean_markdown(text)
                story.append(Paragraph(text, self.styles['CustomBody']))
            
            # 分隔線
            elif line.startswith('---') or line.startswith('***'):
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph('_' * 80, self.styles['CustomBody']))
                story.append(Spacer(1, 0.2*inch))
            
            # 普通段落
            else:
                # 清理 Markdown 格式
                text = self._clean_markdown(line)
                if text:
                    story.append(Paragraph(text, self.styles['CustomBody']))
                    story.append(Spacer(1, 0.1*inch))
        
        return story
    
    def _clean_markdown(self, text: str) -> str:
        """清理 Markdown 格式標記"""
        # 先處理粗體標記（兩個星號或底線）
        text = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
        text = re.sub(r'__(.*?)__', r'<b>\1</b>', text)
        
        # 再處理斜體標記（單個星號或底線，但不能是已經處理過的粗體）
        # 使用負向先行斷言避免匹配 <b> 標籤內的內容
        text = re.sub(r'(?<!\*)\*(?!\*)([^*]+?)(?<!\*)\*(?!\*)', r'<i>\1</i>', text)
        text = re.sub(r'(?<!_)_(?!_)([^_]+?)(?<!_)_(?!_)', r'<i>\1</i>', text)
        
        # 清理連結
        text = self._clean_markdown_links(text)
        
        return text
    
    def _clean_markdown_links(self, text: str) -> str:
        """清理 Markdown 連結格式"""
        # [文字](URL) -> 文字 (URL)
        text = re.sub(r'\[(.*?)\]\((.*?)\)', r'\1 (<font color="blue">\2</font>)', text)
        return text
    
    def generate_excel(
        self, 
        news_data: List[Dict[str, str]], 
        filename: Optional[str] = None
    ) -> Path:
        """
        生成 Excel 報告
        
        Args:
            news_data: 結構化的新聞數據列表
            filename: 可選的文件名，不提供則自動生成
            
        Returns:
            Path: 生成的 Excel 文件路徑
        """
        print("📊 Report Generator Agent 開始生成 Excel...")
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"東南亞金融新聞報告_{timestamp}.xlsx"
        
        # 確保文件名以 .xlsx 結尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        excel_path = self.reports_dir / filename
        
        try:
            # 創建 DataFrame
            df = pd.DataFrame(news_data)
            
            # 調整列順序，加入「摘要」和「重點分析」，移除「關鍵字」和「來源」
            columns_order = ['新聞標題（中文）', '來源國家', '來源網站連結', '發布日期', '摘要', '重點分析']
            existing_columns = [col for col in columns_order if col in df.columns]
            df = df[existing_columns]
            
            # 使用 openpyxl 引擎寫入 Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='新聞報告')
                
                # 獲取工作表並調整列寬
                worksheet = writer.sheets['新聞報告']
                
                # 設置列寬（調整後的順序）
                column_widths = {
                    'A': 50,  # 新聞標題（中文）
                    'B': 15,  # 來源國家
                    'C': 60,  # 來源網站連結
                    'D': 15,  # 發布日期
                    'E': 80,  # 摘要
                    'F': 80,  # 重點分析
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 設置標題列樣式（粗體、置中）
                from openpyxl.styles import Font, Alignment, PatternFill
                
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # 設置資料列樣式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for idx, cell in enumerate(row):
                        cell.alignment = left_alignment
                        # 如果是連結列（C列），設置為藍色字體
                        if idx == 2:  # 來源網站連結
                            cell.font = Font(color='0000FF', underline='single')
            
            print(f"✅ Excel 生成成功: {excel_path}")
            return excel_path
            
        except Exception as e:
            print(f"❌ Excel 生成失敗: {str(e)}")
            raise
    



if __name__ == "__main__":
    # 測試 Report Generator Agent
    agent = ReportGeneratorAgent()
    
    test_markdown = """
# 東南亞金融新聞報告

## 報告摘要
本報告涵蓋東南亞主要國家的金融市場動態。

## 新聞詳情

### 1. 新加坡股市上漲
- **來源**：[Bloomberg](https://bloomberg.com)
- **摘要**：新加坡海峽時報指數今日上漲 1.2%

## 市場洞察
- 區域經濟持續復甦
- 投資信心增強
"""
    
    pdf_path = agent.generate_pdf(test_markdown, "test_report.pdf")
    print(f"測試 PDF 已生成: {pdf_path}")
